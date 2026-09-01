"""
Dropboxの「機材レンタル」フォルダ配下にある月ごとのサブフォルダから、
各月の「○年○月_見積データ抽出.xlsx」を探して取得し、金額情報を除いた
構造化JSON（zaiki_app/assets/data/estimate_YYYY_MM.json）を生成するスクリプト。

このスクリプトはインターネットに出られるパソコン上（VSCodeのターミナルなど）で
実行してください。実行後、zaiki_app/assets/data/ 以下に各月のJSONファイルと
index.json が生成/更新されます。あとはアプリをホットリスタートすれば
「見積抽出」タブに反映されます。

使い方:
    python fetch_estimate_data.py --token "sl.xxxxxxxx" --root "/野添/機材レンタル"

トークンは環境変数 DROPBOX_ACCESS_TOKEN でも指定できます。
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import tempfile
import urllib.error
import urllib.request
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(
        "openpyxl がインストールされていません。\n"
        "  pip install -r requirements-pdf-extract.txt\n"
        "を実行してから再度お試しください。",
        file=sys.stderr,
    )
    raise

API = "https://api.dropboxapi.com/2"
CONTENT_API = "https://content.dropboxapi.com/2"


def _post(endpoint, token, body, extra_headers=None, timeout=30):
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    if extra_headers:
        headers.update(extra_headers)
    req = urllib.request.Request(
        f"{API}/{endpoint}",
        data=json.dumps(body).encode("utf-8"),
        headers=headers,
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"{endpoint} 失敗: HTTP {e.code} {detail[:300]}") from e


def _download(token, path, extra_headers=None, timeout=60):
    headers = {
        "Authorization": f"Bearer {token}",
        "Dropbox-API-Arg": json.dumps({"path": path}),
    }
    if extra_headers:
        headers.update(extra_headers)
    req = urllib.request.Request(
        f"{CONTENT_API}/files/download", headers=headers, method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return resp.read()
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"files/download 失敗: HTTP {e.code} {detail[:300]}") from e


def _get_team_member_id(token):
    """Dropbox Businessアカウントの場合、team/members/list からteam_member_idを推定する。
    個人アカウントの場合は失敗するのでNoneを返す。"""
    try:
        resp = _post("team/members/list", token, {"limit": 1})
    except RuntimeError:
        return None
    members = resp.get("members") or []
    if not members:
        return None
    return members[0].get("profile", {}).get("team_member_id")


def _list_folder_recursive(token, path, extra_headers):
    resp = _post(
        "files/list_folder",
        token,
        {
            "path": path,
            "recursive": True,
            "include_deleted": False,
        },
        extra_headers,
    )
    entries = list(resp.get("entries", []))
    while resp.get("has_more"):
        resp = _post(
            "files/list_folder/continue",
            token,
            {"cursor": resp["cursor"]},
            extra_headers,
        )
        entries.extend(resp.get("entries", []))
    return entries


def _resolve_entries(token, root_path):
    """team_member_idありなし両方を試し、最初に成功した方の一覧を返す。"""
    attempts = []
    team_member_id = _get_team_member_id(token)
    if team_member_id:
        attempts.append({"Dropbox-API-Select-User": team_member_id})
    attempts.append(None)  # 個人アカウント/権限が不要なケース

    last_error = None
    for headers in attempts:
        try:
            entries = _list_folder_recursive(token, root_path, headers)
            return entries, headers
        except RuntimeError as e:
            last_error = e
            continue
    raise RuntimeError(f"フォルダを開けませんでした: {root_path} ({last_error})")


MONTH_RE = re.compile(r"(\d{4})年(\d{1,2})月")
FILENAME_MONTH_RE = re.compile(r"estimate_(\d{4})_(\d{2})\.json")


def _month_window(months_ahead, today=None):
    """今月を含めて months_ahead か月先までの (year, month) の集合を返す。"""
    today = today or date.today()
    window = set()
    y, m = today.year, today.month
    for i in range(months_ahead + 1):
        total = (y * 12 + (m - 1)) + i
        window.add((total // 12, total % 12 + 1))
    return window


def _is_blank(row):
    return all(c is None for c in row)


def _is_discount_row(kind, name):
    text = f"{kind or ''}{name or ''}"
    return ("割引" in text) or ("値引" in text) or ("協力金" in text)


def parse_estimate_workbook(xlsx_path):
    """見積データ抽出.xlsx（見積シート＋サマリーシート）を金額を除いて構造化する。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["見積"]
    rows = list(ws.iter_rows(min_row=5, values_only=True))

    jobs = []
    current = None
    current_subcat = None

    for row in rows:
        if _is_blank(row):
            continue
        kind, name, memo, unit_price, qty, duration, amount = row
        if kind == "小計":
            current = None
            current_subcat = None
            continue
        if kind is not None and all(c is None for c in row[1:]):
            if str(kind).startswith("▼"):
                current_subcat = str(kind).replace("▼", "").strip()
                continue
            else:
                current = {"folder": str(kind).strip(), "items": []}
                jobs.append(current)
                current_subcat = None
                continue
        if current is None:
            continue
        if _is_discount_row(kind, name):
            continue
        item = {
            "category": current_subcat or kind,
            "name": name,
            "memo": memo,
            "qty": qty,
            "duration": duration,
        }
        item = {k: v for k, v in item.items() if v not in (None, "")}
        if item.get("name"):
            current["items"].append(item)

    ws2 = wb["サマリー"]
    summary_rows = [
        r for r in ws2.iter_rows(min_row=5, values_only=True) if not _is_blank(r) and r[1] is not None
    ]

    result = []
    for j, s in zip(jobs, summary_rows):
        (
            _,
            folder_sum,
            oname,
            address,
            address_detail,
            ship_date,
            ship_time,
            return_time,
        ) = (
            s[0],
            s[1],
            s[2],
            s[3],
            s[4],
            s[5],
            s[6],
            s[7],
        )
        note = s[12] if len(s) > 12 else None
        entry = {"folder": j["folder"]}
        if oname:
            entry["clientName"] = oname
        if address and address != "-":
            entry["deliveryAddress"] = address
        if address_detail and address_detail != "-":
            entry["deliveryAddressDetail"] = address_detail
        if ship_date:
            entry["deliveryDate"] = str(ship_date)
        if ship_time and ship_time != "-":
            entry["time"] = ship_time
        if return_time and return_time != "-":
            entry["returnTime"] = return_time
        if note:
            entry["note"] = note
        entry["items"] = j["items"]
        result.append(entry)

    return result


def _write_outputs(files, out_dir):
    """files: (year, month, name, xlsx_path, source_label) のリスト。戻り値: 生成したファイル名のリスト。"""
    generated = []
    for year, month, name, xlsx_path, source_label in files:
        month_key = f"{year:04d}_{month:02d}"
        try:
            jobs = parse_estimate_workbook(xlsx_path)
        except Exception as e:
            print(f"解析に失敗しました({source_label}): {e}", file=sys.stderr)
            continue

        out = {
            "month": f"{year:04d}-{month:02d}",
            "source": f"{source_label}（{name}）",
            "note": "金額・単価・小計等の金銭情報は含みません",
            "jobs": jobs,
        }
        out_file = out_dir / f"estimate_{month_key}.json"
        out_file.write_text(
            json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(f"書き出し: {out_file} ({len(jobs)}件)")
        generated.append(out_file.name)
    return generated


def _fetch_local(local_root):
    """Dropboxデスクトップアプリの同期フォルダから直接読み込む。"""
    root = Path(local_root)
    if not root.exists():
        print(f"ローカルフォルダが見つかりません: {root}", file=sys.stderr)
        return []

    candidates = [
        p
        for p in root.rglob("*")
        if p.is_file()
        and "見積データ抽出" in p.name
        and p.suffix.lower() in (".xlsx", ".xlsm")
    ]
    print(f"対象ファイル数: {len(candidates)}")

    files = []
    for p in candidates:
        print(f"  - {p}")
        m = MONTH_RE.search(p.name) or MONTH_RE.search(str(p))
        if not m:
            print(f"月が判定できずスキップ: {p}")
            continue
        year, month = int(m.group(1)), int(m.group(2))
        files.append((year, month, p.name, p, "Dropbox見積データ抽出（ローカル同期）"))
    return files


def _fetch_dropbox_api(token, dropbox_root):
    print(f"Dropboxフォルダを検索中: {dropbox_root}")
    entries, headers = _resolve_entries(token, dropbox_root)
    print(f"エントリ数: {len(entries)}")

    target_files = [
        e
        for e in entries
        if e.get(".tag") == "file"
        and "見積データ抽出" in (e.get("name") or "")
        and str(e.get("name", "")).lower().endswith((".xlsx", ".xlsm"))
    ]
    print(f"対象ファイル数: {len(target_files)}")
    for f in target_files:
        print(f"  - {f.get('path_display')}")

    files = []
    tmp_dir = tempfile.mkdtemp()
    for f in target_files:
        path_display = f.get("path_display") or ""
        name = f.get("name") or ""
        m = MONTH_RE.search(name) or MONTH_RE.search(path_display)
        if not m:
            print(f"月が判定できずスキップ: {path_display}")
            continue
        year, month = int(m.group(1)), int(m.group(2))

        print(f"ダウンロード中: {path_display}")
        data = _download(token, path_display, headers)
        tmp_path = Path(tmp_dir) / name
        tmp_path.write_bytes(data)
        files.append((year, month, name, tmp_path, "Dropbox見積データ抽出"))
    return files


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--token",
        default=os.getenv("DROPBOX_ACCESS_TOKEN", ""),
        help="Dropboxアクセストークン。省略時は環境変数 DROPBOX_ACCESS_TOKEN を使用（--local-root指定時は不要）",
    )
    parser.add_argument(
        "--root",
        default="/野添/機材レンタル",
        help="見積データ抽出.xlsxを探すDropbox上のルートフォルダ（デフォルト: /野添/機材レンタル）",
    )
    parser.add_argument(
        "--local-root",
        default=os.getenv(
            "ESTIMATE_LOCAL_ROOT",
            r"C:\Users\246pa\246GROUP Dropbox\246 dropbox\野添\機材レンタル",
        ),
        help="Dropboxデスクトップアプリの同期フォルダから直接読み込む場合のパス。"
        "指定時はDropbox APIを使わない。空文字を指定するとこの機能を無効化する。",
    )
    parser.add_argument(
        "--out",
        default=str(Path(__file__).resolve().parent.parent / "assets" / "data"),
        help="出力先ディレクトリ（デフォルト: ../assets/data）",
    )
    parser.add_argument(
        "--months-ahead",
        type=int,
        default=3,
        help="今月を含めて何か月先まで対象にするか（デフォルト: 3＝今月+3か月先の計4か月分）",
    )
    args = parser.parse_args()

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    window = _month_window(args.months_ahead)

    if args.local_root:
        print(f"ローカル同期フォルダを検索中: {args.local_root}")
        files = _fetch_local(args.local_root)
    else:
        token = args.token.strip()
        if not token:
            print("Dropboxアクセストークンが指定されていません。--token または環境変数 DROPBOX_ACCESS_TOKEN で指定してください。", file=sys.stderr)
            return 1
        files = _fetch_dropbox_api(token, args.root)

    skipped = [f for f in files if (f[0], f[1]) not in window]
    for year, month, name, _path, _label in skipped:
        print(f"対象期間外のためスキップ: {year:04d}年{month:02d}月 ({name})")
    files = [f for f in files if (f[0], f[1]) in window]

    generated = _write_outputs(files, out_dir)

    # index.json を対象期間（今月+--months-ahead か月先）の月だけに絞り込む
    index_path = out_dir / "index.json"
    existing = []
    if index_path.exists():
        try:
            existing = json.loads(index_path.read_text(encoding="utf-8")).get("files", [])
        except Exception:
            existing = []

    def _in_window(filename):
        m = FILENAME_MONTH_RE.match(filename)
        return bool(m) and (int(m.group(1)), int(m.group(2))) in window

    all_files = sorted(
        f for f in (set(existing) | set(generated)) if _in_window(f)
    )
    index_path.write_text(
        json.dumps({"files": all_files}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"index.json 更新: {all_files}")

    if not generated:
        print("新規に取得できたファイルはありませんでした。", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
